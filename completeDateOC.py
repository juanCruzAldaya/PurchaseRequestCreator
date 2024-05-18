from openpyxl import load_workbook
from openpyxl import Workbook
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
import time
from datetime import datetime, timedelta



def get_date_30_days_before():
    # Get the current date
    current_date = datetime.now()

    # Calculate the date 30 days before
    previous_date = current_date - timedelta(days=60)

    # Format the date as day/month/year
    formatted_date = previous_date.strftime("%d/%m/%Y")

    return formatted_date

def find_row_with_value_in_column(sheet, column_letter, search_value, max_row):
    try:
        # Iterate through the specified column up to the given max_row
        for row_num in range(1, max_row + 1):
            cell_value = sheet[f"{column_letter}{row_num}"].value
            if cell_value == search_value:
                return row_num

        # If not found, return None
        return None

    except KeyError:
        print(f"Sheet '{sheet}' not found in the workbook.")
        return None


def convert_date_format(date_str):
    # Spanish month abbreviations to month number mapping
    months = {
        'ene': '1', 'feb': '2', 'mar': '3', 'abr': '4',
        'may': '5', 'jun': '6', 'jul': '7', 'ago': '8',
        'sep': '9', 'oct': '10', 'nov': '11', 'dic': '12'
    }
    
    # Split the date string into day, month abbreviation, and year
    day, month_abbr, year = date_str.split()
    
    # Replace the Spanish month abbreviation with the month number
    month = months[month_abbr.lower()]
    
    # Return the formatted date string
    return f"{month}/{day}/{year}"


option = EdgeOptions()

option.add_argument("start-maximized")

driver = webdriver.Edge(options = option)

actions = ActionChains(driver)

timeOut = 40

workbook = load_workbook(filename='OCSinFecha.xlsx')
sheet = workbook.active
max_row = sheet.max_row

i = 2

driver.get("https://s1.ariba.com/Buyer/Main/aw?awh=r&awssk=Be2NoZxjNEANOO0p&realm=accenture&awrdt=1")
try: 
    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/table/tbody/tr/td[1]/div/div/ul/span[1]/li/span/a')))
    inicioButton = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/table/tbody/tr/td[1]/div/div/ul/span[1]/li/span/a')
    inicioButton.click()
except TimeoutException:
    print ("SAP failed, run script again...")


time.sleep(1)

try: 
    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td[2]/div/form/table/tbody/tr/td[3]/input')))
    idOc = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td[2]/div/form/table/tbody/tr/td[3]/input')
    idOc.send_keys(sheet["B"+str(i)].value)

except TimeoutException:
    print ("SAP failed, run script again...")


try: 
    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td[2]/div/form/table/tbody/tr/td[5]/button')))
    searchButton = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td[2]/div/form/table/tbody/tr/td[5]/button')
    searchButton.click()
except TimeoutException:
    print ("SAP failed, run script again...")
try: 
    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr/td/div[1]/table[2]/tbody/tr[1]/td[3]/div/table[2]/tbody/tr[9]/td[2]/a[2]')))
    responsibleButton = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr/td/div[1]/table[2]/tbody/tr[1]/td[3]/div/table[2]/tbody/tr[9]/td[2]/a[2]')
    responsibleButton.click()
except TimeoutException:
    print ("SAP failed, run script again...")


try: 
    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/div[2]/div/a[2]')))
    suprButton = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/div[2]/div/a[2]')
    suprButton.click()
except TimeoutException:
    print ("SAP failed, run script again...")
time.sleep(1)


try: 
    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[2]/tbody/tr[1]/td[1]/div/table[2]/tbody/tr[8]/td[2]/table/tbody/tr[2]/td[2]/nobr/div[2]/div[1]/input')))
    sinceSearch = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[2]/tbody/tr[1]/td[1]/div/table[2]/tbody/tr[8]/td[2]/table/tbody/tr[2]/td[2]/nobr/div[2]/div[1]/input')
    sinceSearch.clear()
    time.sleep(1)
    sinceSearch.send_keys(str(get_date_30_days_before()))
    
except TimeoutException:
    print ("SAP failed, run script again...")

try: 
    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[2]/td/div[1]/table/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[4]/table/tbody/tr/td[1]/span/a/div')))
    filterButton = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[2]/td/div[1]/table/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[4]/table/tbody/tr/td[1]/span/a/div')
    filterButton.click()
except TimeoutException:
    print ("SAP failed, run script again...")



time.sleep(1)
try: 
    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/div[2]/div/a[8]')))
    sendingDate = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/div[2]/div/a[8]')
    sendingDate .click()
except TimeoutException:
    print ("SAP failed, run script again...")
time.sleep(2)


try:
    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[3]/tbody/tr/td[2]/table/tbody/tr/td[1]/button')))
    searchButton = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[3]/tbody/tr/td[2]/table/tbody/tr/td[1]/button')
    searchButton.click()
except TimeoutException:
    print ("SAP failed, run script again...")
time.sleep(2)

driver.execute_script("window.scrollTo(5,document.body.scrollHeight)")
driver.execute_script("window.scrollTo(5,document.body.scrollHeight)")

try: 
    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[2]/td/div[1]/table/tbody/tr[2]/td/table/tbody/tr[2]/td/div/table/tbody/tr[2]/td[8]/span/table/tbody/tr/td')))
    searchButton = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[2]/td/div[1]/table/tbody/tr[2]/td/table/tbody/tr[2]/td/div/table/tbody/tr[2]/td[8]/span/table/tbody/tr/td')
    sheet["C"+str(i)] = convert_date_format(searchButton.text)
except TimeoutException:
    print ("SAP failed, run script again...")
time.sleep(2)

workbook.save(filename='OCSinFecha.xlsx')


while True:
    i += 1
    if sheet["A"+str(i)].value != None:
        try: 
            WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[2]/tbody/tr[1]/td[3]/div/table[2]/tbody/tr[1]/td[2]/input')))
            ocField = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[2]/tbody/tr[1]/td[3]/div/table[2]/tbody/tr[1]/td[2]/input')
            ocField.clear()
            ocField.send_keys(sheet["B"+str(i)].value)

        except TimeoutException:
            print ("SAP failed, run script again...")
        time.sleep(1)





        try: 
            WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[3]/tbody/tr/td[2]/table/tbody/tr/td[1]/button')))
            searchButton = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[3]/tbody/tr/td[2]/table/tbody/tr/td[1]/button')
            searchButton.click()
        except TimeoutException:
            print ("SAP failed, run script again...")
        time.sleep(2)        

        try: 
            WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[2]/td/div[1]/table/tbody/tr[2]/td/table/tbody/tr[2]/td/div/table/tbody/tr[2]/td[6]/span/table/tbody/tr/td')))
            dateOC = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[2]/td/div[1]/table/tbody/tr[2]/td/table/tbody/tr[2]/td/div/table/tbody/tr[2]/td[6]/span/table/tbody/tr/td')
            sheet["D"+str(i)] = convert_date_format(dateOC.text)
        except TimeoutException:
            print ("SAP failed, run script again...")
        time.sleep(1)

        workbook.save(filename='OCSinFecha.xlsx')

    else:
        break
    