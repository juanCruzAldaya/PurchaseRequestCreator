from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
import time
from datetime import datetime, timedelta
import calendar
import os



def get_next_weekday_one_month_later():
    # Get the current date
    today = datetime.today()

    # Calculate the date one month later
    one_month_later = today + timedelta(days=30)

    # Check if the date is a weekend (Saturday or Sunday)
    while one_month_later.weekday() >= 5:
        one_month_later += timedelta(days=1)

    return one_month_later.strftime('%d/%m/%Y')



def substring_before_at(text):
    # Find the position of the '@' character
    at_position = text.find('@')
    # If '@' is found, return the substring up to '@'
    if at_position != -1:
        return text[:at_position]
    # If '@' is not found, return the original string
    return text
def element_exists(driver, xpath):
    try:
        driver.find_element(By.XPATH, xpath)
        return True
    except NoSuchElementException:
        return False

option = EdgeOptions()

option.add_argument("start-maximized")

driver = webdriver.Edge(options = option)

actions = ActionChains(driver)

timeOut = 40

workbook = load_workbook(filename = "PRSinHacerFS.xlsx") ###
sheet = workbook.active
max_row = sheet.max_row
i = 1

for row in sheet.iter_rows(min_row=2, min_col=1, max_row=max_row, max_col=1):
    if (sheet["A"+str(i)].value != None) and (sheet["l"+str(i+1)].value == None):
        driver.get("edge://settings/clearBrowserData")
        driver.get("edge://settings/clearBrowserData")
        time.sleep(2) 

        try: 
            WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div/div/div/div[2]/div/div[2]/div/div[2]/button[1]')))
            clearCache = driver.find_element(By.XPATH, '/html/body/div[2]/div/div/div/div[2]/div/div[2]/div/div[2]/button[1]')
            clearCache.click()
        except TimeoutException:
            print ("SAP failed, run script again...")

        
        driver.get("https://s1.ariba.com/Buyer/Main/aw?awh=r&awssk=5Z3OrytN1OHAwAhu&realm=accenture&awrdt=1") 
        os.system("cls")
        for cell in row:
            cellRow = str(cell.row)
            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/table/tbody/tr/td[3]/div/table/tbody/tr/td[3]/a/div')))
                newPR = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/table/tbody/tr/td[3]/div/table/tbody/tr/td[3]/a/div')
                newPR.click()


            except TimeoutException:
                print ("SAP failed, run script again...")

            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/div/div/a[3]')))
                createPR = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/div/div/a[3]')
                createPR.click()
            except TimeoutException:
                print ("SAP failed, run script again...")


            if sheet["i"+cellRow].value != 0: ## si tiene precio le agrega 'precio de lista' al titulo, si no nada

                try: 
                    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/table[1]/tbody/tr/td[1]/table[1]/tbody/tr[2]/td[3]/table/tbody/tr/td[1]/input')))
                    titleField = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/table[1]/tbody/tr/td[1]/table[1]/tbody/tr[2]/td[3]/table/tbody/tr/td[1]/input')
                    titleField.send_keys("ARG_LTS_" + str(sheet["h" + cellRow].value) + "_" + str(sheet["e"+cellRow].value) + "_" +  str(sheet["b"+cellRow].value) + "_" + str(sheet["D"+cellRow].value) + " (Precio de lista)")
                except TimeoutException:
                    print ("SAP failed, run script again...")
            else:

                try: 
                    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/table[1]/tbody/tr/td[1]/table[1]/tbody/tr[2]/td[3]/table/tbody/tr/td[1]/input')))
                    titleField = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/table[1]/tbody/tr/td[1]/table[1]/tbody/tr[2]/td[3]/table/tbody/tr/td[1]/input')
                    titleField.send_keys("ARG_LTS_" + str(sheet["h" + cellRow].value) + "_" + str(sheet["e"+cellRow].value) + "_" +  str(sheet["b"+cellRow].value) + "_" + str(sheet["D"+cellRow].value))
                except TimeoutException:
                    print ("SAP failed, run script again...")


            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/table[1]/tbody/tr/td[1]/table[1]/tbody/tr[3]/td[3]/div/div/span/div[1]/div[1]/input')))
                inNameOf = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/table[1]/tbody/tr/td[1]/table[1]/tbody/tr[3]/td[3]/div/div/span/div[1]/div[1]/input')
                inNameOf.clear()
                inNameOf.send_keys("Andres J. Renteria")

            except TimeoutException:
                print ("SAP failed, run script again...")
            time.sleep(2)


            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/table[1]/tbody/tr/td[1]/table[1]/tbody/tr[4]/td[3]/div/div/span/div[1]/div[1]/input')))
                companyCode = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/table[1]/tbody/tr/td[1]/table[1]/tbody/tr[4]/td[3]/div/div/span/div[1]/div[1]/input')
                companyCode.clear()
                companyCode.send_keys("1300")
                
            except TimeoutException:
                print ("SAP failed, run script again...")

            time.sleep(3)
            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/table[1]/tbody/tr/td[1]/table[1]/tbody/tr[7]/td[3]/table/tbody/tr/td[1]/div/div/span/div[1]/div[1]/input')))
                businessApprover = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/table[1]/tbody/tr/td[1]/table[1]/tbody/tr[7]/td[3]/table/tbody/tr/td[1]/div/div/span/div[1]/div[1]/input')
                businessApprover.clear()
                businessApprover.send_keys("Facundo N.  Petit")
            except TimeoutException:
                print ("SAP failed, run script again...")

            time.sleep(1)

            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/table[1]/tbody/tr/td[1]/div[2]/span/div/div[1]/table/tbody/tr[3]/td[3]/table/tbody/tr/td[1]/nobr/div[2]/div[1]/input')))
                dateToRecive = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/table[1]/tbody/tr/td[1]/div[2]/span/div/div[1]/table/tbody/tr[3]/td[3]/table/tbody/tr/td[1]/nobr/div[2]/div[1]/input')
                dateToRecive.clear()
                dateToRecive.send_keys(str(get_next_weekday_one_month_later()))

            except TimeoutException:
                print ("SAP failed, run script again...")

            time.sleep(2)


            try:
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[1]/td/table/tbody/tr/td[2]/table/tbody/tr/td[2]/div/button')))
                continueBuy = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[1]/td/table/tbody/tr/td[2]/table/tbody/tr/td[2]/div/button')
                continueBuy.click()

            except TimeoutException:
                print ("SAP failed, run script again...")


            time.sleep(3)
            #
            try:
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div/div[3]/table/tbody/tr/td/div/table/tbody/tr/td[2]/form/div[2]/table/tbody/tr/td[3]/div/button')))
                addArticle = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div/div[3]/table/tbody/tr/td/div/table/tbody/tr/td[2]/form/div[2]/table/tbody/tr/td[3]/div/button')
                addArticle.click()

            except TimeoutException:
                print ("SAP failed, run script again...")
            time.sleep(2)

            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr[2]/td[3]/div/div/span/div[1]/div[1]/input')))
                organizationCode = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr[2]/td[3]/div/div/span/div[1]/div[1]/input')
                organizationCode.clear()
                organizationCode.send_keys("1300")

            except TimeoutException:
                print ("SAP failed, run script again...")
            time.sleep(2)

            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr[2]/td[3]/div/div/span/div[1]/div[1]/input')))
                organizationCode = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr[2]/td[3]/div/div/span/div[1]/div[1]/input')
                organizationCode.clear()
                organizationCode.send_keys("1300")

            except TimeoutException:
                print ("SAP failed, run script again...")
            time.sleep(2)

            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div[1]/table/tbody/tr/td[1]/table/tbody/tr[2]/td[3]/textarea')))
                itemDesc = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div[1]/table/tbody/tr/td[1]/table/tbody/tr[2]/td[3]/textarea')
                itemDesc.clear()
                itemDesc.send_keys ("Parte: " + str(sheet["e"+cellRow].value) + " de " + " "+ str(sheet["b"+cellRow].value) + " " + str(sheet["d"+cellRow].value) + " \nPedida por: " + substring_before_at(str(sheet["c"+cellRow].value)) + "\nPart Number | FRU: " + str(sheet["g"+cellRow].value)) 

            except TimeoutException:
                print ("SAP failed, run script again...")
            time.sleep(1)

            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div[1]/table/tbody/tr/td[1]/table/tbody/tr[4]/td[3]/input')))
                itemQuantity = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div[1]/table/tbody/tr/td[1]/table/tbody/tr[4]/td[3]/input')
                itemQuantity.clear()
                quant = sheet["f"+cellRow].value
                itemQuantity.send_keys (str(quant)) 
                time.sleep(1)
                actions.send_keys(Keys.TAB).perform()
            except TimeoutException:
                print ("SAP failed, run script again...")

            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div[1]/table/tbody/tr/td[1]/table/tbody/tr[3]/td[3]/div/div/span/div[1]/div[1]/input')))
                mercCode = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div[1]/table/tbody/tr/td[1]/table/tbody/tr[3]/td[3]/div/div/span/div[1]/div[1]/input')
                mercCode.send_keys ("Computer Hardware: Computer accessories, docking station, port, speaker, USB, switch, adapter")
                time.sleep(1)
                actions.send_keys(Keys.TAB).perform()
            except TimeoutException:
                print ("SAP failed, run script again...")

            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div[1]/table/tbody/tr/td[1]/table/tbody/tr[7]/td[3]/span/input')))
                itemPrice = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div[1]/table/tbody/tr/td[1]/table/tbody/tr[7]/td[3]/span/input')
                itemPrice.clear()
                pricePerUnite = sheet["i"+cellRow].value
                totlalPrice = quant * pricePerUnite
                itemPrice.send_keys (str(totlalPrice))
                time.sleep(1)
                actions.send_keys(Keys.TAB).perform()
            except TimeoutException:
                print ("SAP failed, run script again...")
            time.sleep(2)
            driver.execute_script("window.scrollTo(5,document.body.scrollHeight)")




            if sheet["j"+cellRow].value != None:
                try: 
                    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div[4]/table/tbody/tr/td[1]/table/tbody/tr[2]/td[3]/div/div/span/div[1]/div[1]/input')))
                    seller = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div[4]/table/tbody/tr/td[1]/table/tbody/tr[2]/td[3]/div/div/span/div[1]/div[1]/input')
                    
                    seller.send_keys (str(sheet["j"+cellRow].value)) 

                except TimeoutException:
                    print ("SAP failed, run script again...")

            else:
                try: 
                    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div[4]/table/tbody/tr/td[1]/table/tbody/tr[2]/td[3]/div/div/span/div[1]/div[1]/input')))
                    seller = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div[4]/table/tbody/tr/td[1]/table/tbody/tr[2]/td[3]/div/div/span/div[1]/div[1]/input')
                    
                    seller.send_keys (" ")

                except TimeoutException:
                    print ("SAP failed, run script again...")

        
                   
            time.sleep(1)      
            # try: 
            #     WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div[4]/table/tbody/tr/td[1]/table/tbody/tr[4]/td[3]/input')))
            #     partNumber = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div[4]/table/tbody/tr/td[1]/table/tbody/tr[4]/td[3]/input')
            #     partNumber.send_keys (str(sheet["f"+cellRow].value)) 

            # except TimeoutException:
            #     print ("SAP failed, run script again...")
            # time.sleep(3)
            
            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr/td[1]/div/button')))                            
                addItemButton = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr/td[1]/div/button')
                addItemButton.click()

            except TimeoutException:
                print ("SAP failed, run script again...")
            
            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div/div[2]/div/form/div[2]/div/div[2]/div[5]/table/tbody/tr/td[2]/a')))
                confirmBuyButton = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div/div[2]/div/form/div[2]/div/div[2]/div[5]/table/tbody/tr/td[2]/a')
                confirmBuyButton.click()

            except TimeoutException:
                print ("SAP failed, run script again...")
            time.sleep(1)


            driver.execute_script("window.scrollTo(5,document.body.scrollHeight)")
            driver.execute_script("window.scrollTo(5,document.body.scrollHeight)")

            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/div[2]/div[1]/div/div[1]/table/tbody/tr[3]/td/table/tbody/tr/td[1]/table/tbody/tr/td[2]/div/table/tbody/tr/td[1]/div/div/button')))
                actionButton = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/div[2]/div[1]/div/div[1]/table/tbody/tr[3]/td/table/tbody/tr/td[1]/table/tbody/tr/td[2]/div/table/tbody/tr/td[1]/div/div/button')
                actionButton.click()

            except TimeoutException:
                print ("SAP failed, run script again...")
            time.sleep(1)

            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/div/div/a[2]')))
                editAction = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/div/div/a[2]')
                editAction.click()

            except TimeoutException:
                print ("SAP failed, run script again...")
            time.sleep(2)

            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[1]/td/table/tbody/tr/td[1]/span[1]')))
                organizationCode = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[1]/td/table/tbody/tr/td[1]/span[1]')
                prNumber = organizationCode.text
                sheet["l"+cellRow].value = prNumber[:-1]
                workbook.save(filename="PRSinHacerFS.xlsx")
            except TimeoutException:
                print ("SAP failed, run script again...")
            #time.sleep(1)            
            driver.execute_script("window.scrollTo(5,150)")
            #
            
            element1_exists = element_exists(driver, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[15]/td[3]/div/div/span/div[1]/div[1]/input')
            element2_exists = element_exists(driver, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[14]/td[3]/div/div/span/div[1]/div[1]/input')

            
            if element2_exists:
                try: 
                    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[14]/td[3]/div/div/span/div[1]/div[1]/input')))
                    organizationCode = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[14]/td[3]/div/div/span/div[1]/div[1]/input')
                    organizationCode.clear()
                    organizationCode.send_keys("1300")

                except TimeoutException:
                    print ("SAP failed, run script again...")
            #time.sleep(1)
            elif element1_exists:
                try:
                    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[15]/td[3]/div/div/span/div[1]/div[1]/input')))
                    organizationCode = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[15]/td[3]/div/div/span/div[1]/div[1]/input')
                    organizationCode.clear()
                    organizationCode.send_keys("1300")

                except TimeoutException:
                    print ("SAP failed, run script again...")
                    
            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/div[4]/table[2]/tbody/tr[2]/td[3]/table/tbody/tr/td[1]/div/div/span/div[1]/div[1]/input')))
                organizationCode = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/div[4]/table[2]/tbody/tr[2]/td[3]/table/tbody/tr/td[1]/div/div/span/div[1]/div[1]/input')
                organizationCode.clear()
                organizationCode.send_keys("B6C0R001")

            except TimeoutException:
                print ("SAP failed, run script again...")

            time.sleep(2)

            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr/td[1]/div/button')))
                acceptButton = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr/td[1]/div/button')
                acceptButton.click()

            except TimeoutException:
                print ("SAP failed, run script again...")
            time.sleep(1)
            driver.execute_script("window.scrollTo(0, 0)")


            time.sleep(1)

            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/table[1]/tbody/tr/td[1]/table[1]/tbody/tr[5]/td[3]/table/tbody/tr/td[1]/div/div/span/div[1]/div[1]/input')))
                organizationCode = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[3]/td/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/table[1]/tbody/tr/td[1]/table[1]/tbody/tr[5]/td[3]/table/tbody/tr/td[1]/div/div/span/div[1]/div[1]/input')
                organizationCode.clear()
                organizationCode.send_keys("Leandro M. Garibaldi")
                time.sleep(1)

            except TimeoutException:
                print ("SAP failed, run script again...")

            time.sleep(1)

            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[1]/td/table/tbody/tr/td[2]/table/tbody/tr/td[5]/div/button')))
                organizationCode = driver.find_element(By.XPATH, '/html/body/div[5]/form/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td/div/div/table/tbody/tr[1]/td/table/tbody/tr/td[2]/table/tbody/tr/td[5]/div/button')
                organizationCode.click()
                time.sleep(10)

            except TimeoutException:
                print ("SAP failed, run script again...")
            nextRow = int(cellRow) + 1
            if sheet["A"+ str(nextRow)] != None:
                original_window = driver.current_window_handle
                driver.switch_to.new_window("window")
                newWindow = driver.current_window_handle
                driver.switch_to.window(original_window)
                driver.close()
                driver.switch_to.window(newWindow)



    i += 1