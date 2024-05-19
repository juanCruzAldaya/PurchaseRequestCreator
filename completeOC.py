from openpyxl import load_workbook
from openpyxl import Workbook
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
import time
from datetime import datetime, timedelta



def get_date_30_days_before():
    # Get the current date
    current_date = datetime.now()

    # Calculate the date 30 days before
    previous_date = current_date - timedelta(days=30)

    # Format the date as day/month/year
    formatted_date = previous_date.strftime("%d/%m/%Y")

    return formatted_date


def convert_date_with_map(date_str):
    # Map of Spanish month abbreviations to month numbers
    months = {
        'ene,': '01', 'feb,': '02', 'mar,': '03', 'abr,': '04',
        'may,': '05', 'jun,': '06', 'jul,': '07', 'ago,': '08',
        'sep,': '09', 'oct,': '10', 'nov,': '11', 'dic,': '12'
    }
    
    # Split the input string into components
    parts = date_str.split()
    
    # Extract the day and year
    day = parts[1]
    year = parts[3]
    
    # Use the map to get the month number
    month = months[parts[2].lower()]
    
    # Format the date as mm/dd/yyyy
    return f"{month}/{day}/{year}"

def find_row_with_value_in_column(sheet, column_letter, search_value, max_row):
    try:
        # Iterate through the specified column up to the given max_row
        for row_num in range(1, max_row + 1):
            cell_value = sheet[f"{column_letter}{row_num}"].value
            if cell_value == search_value:
                return row_num

        # If not found, return None
        return None

    except KeyError:
        print(f"Sheet '{sheet}' not found in the workbook.")
        return None

option = EdgeOptions()

option.add_argument("start-maximized")

driver = webdriver.Edge(options = option)

actions = ActionChains(driver)

timeOut = 40

workbook = load_workbook(filename='PRPedidos.xlsx')
sheet = workbook.active
max_row = sheet.max_row

i = 2

driver.get("################################################")
try: 
    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/table/tbody/tr/td[1]/div/div/ul/span[1]/li/span/a')))
    inicioButton = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/table/tbody/tr/td[1]/div/div/ul/span[1]/li/span/a')
    inicioButton.click()
except TimeoutException:
    print ("SAP failed, run script again...")


time.sleep(1)

try: 
    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[3]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr/td[2]/div[2]/table/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td/div[1]/div/div[8]/div/div/div/div[2]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/a')))
    seeMoreButton = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[3]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr/td[2]/div[2]/table/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td/div[1]/div/div[8]/div/div/div/div[2]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/a')
    seeMoreButton.click()
except TimeoutException:
    print ("SAP failed, run script again...")


try: 
    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/div[2]/div/a[1]')))
    ocButton = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/div[2]/div/a[1]')
    ocButton.click()
except TimeoutException:
    print ("SAP failed, run script again...")


try: 
    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[2]/tbody/tr[1]/td[1]/div/table[2]/tbody/tr[18]/td[2]/table/tbody/tr[2]/td[2]/nobr/div[2]/div[1]/input')))
    sinceSearch = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[2]/tbody/tr[1]/td[1]/div/table[2]/tbody/tr[18]/td[2]/table/tbody/tr[2]/td[2]/nobr/div[2]/div[1]/input')
    sinceSearch.clear()
    time.sleep(1)
    sinceSearch.send_keys(str(get_date_30_days_before()))
    
except TimeoutException:
    print ("SAP failed, run script again...")

time.sleep(2)  
driver.execute_script("window.scrollTo(5,document.body.scrollHeight)")
driver.execute_script("window.scrollTo(5,document.body.scrollHeight)")
driver.execute_script("window.scrollTo(5,document.body.scrollHeight)")
driver.execute_script("window.scrollTo(5,document.body.scrollHeight)")

try: 
    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[2]/tbody/tr[1]/td[1]/div/table[2]/tbody/tr[15]/td[2]/div/div')))
    stateOptions = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[2]/tbody/tr[1]/td[1]/div/table[2]/tbody/tr[15]/td[2]/div/div')
    stateOptions.click()
except TimeoutException:
    print ("SAP failed, run script again...")


try: 
    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[2]/tbody/tr[1]/td[1]/div/table[2]/tbody/tr[15]/td[2]/div/div/div/div[10]')))
    pedidoButton = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[2]/tbody/tr[1]/td[1]/div/table[2]/tbody/tr[15]/td[2]/div/div/div/div[10]')
    pedidoButton.click()
except TimeoutException:
    print ("SAP failed, run script again...")


try: 
    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[3]/tbody/tr/td[2]/table/tbody/tr/td[1]/button')))
    searchButton = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[3]/tbody/tr/td[2]/table/tbody/tr/td[1]/button')
    searchButton.click()
except TimeoutException:
    print ("SAP failed, run script again...")
time.sleep(2)



try: 
    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[2]/td/div[1]/table/tbody/tr[2]/td/table/tbody/tr[2]/td/div/table/tbody/tr['+str(i)+']/td[2]')))
    searchButton = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[2]/td/div[1]/table/tbody/tr[2]/td/table/tbody/tr[2]/td/div/table/tbody/tr['+str(i)+']/td[2]')
    something = searchButton.text
except TimeoutException:
    print ("SAP failed, run script again...")

time.sleep(2)
while True:
    try:
        WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[2]/td/div[1]/table/tbody/tr[2]/td/table/tbody/tr[2]/td/div/table/tbody/tr['+str(i)+']/td[2]')))
        prField = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[2]/td/div[1]/table/tbody/tr[2]/td/table/tbody/tr[2]/td/div/table/tbody/tr['+str(i)+']/td[2]')
        prValue = prField.text
        
        stateField = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[2]/td/div[1]/table/tbody/tr[2]/td/table/tbody/tr[2]/td/div/table/tbody/tr['+str(i)+']/td[4]')
        stateValue = stateField.text
        
        rowNumber = find_row_with_value_in_column(sheet, 'A', prValue, max_row)
        
        if (rowNumber != None) and (stateValue == 'Pedido'):
            prField.click()
            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[1]/form/table[1]/tbody/tr[4]/td/div/div[2]/div/div[2]/div[1]/div/div[1]/table/tbody/tr[2]/td/table/tbody/tr[2]/td[4]/span/table/tbody/tr/td/a/span')))
                ocField = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[1]/form/table[1]/tbody/tr[4]/td/div/div[2]/div/div[2]/div[1]/div/div[1]/table/tbody/tr[2]/td/table/tbody/tr[2]/td[4]/span/table/tbody/tr/td/a/span')
                ocValue = ocField.text
            except TimeoutException:
                print ("SAP failed, run script again...")
            
            time.sleep(1)
            
            sheet["C"+str(rowNumber)] = ocValue
            workbook.save("############")

            time.sleep(1)
            ocField.click()

            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div/form/table[1]/tbody/tr[4]/td/div/div[1]/ul/li[6]/span/a')))
                historyField = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div/form/table[1]/tbody/tr[4]/td/div/div[1]/ul/li[6]/span/a')
                historyField.click()
            except TimeoutException:
                print ("SAP failed, run script again...")
            
            time.sleep(1)                    
            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div/form/table[1]/tbody/tr[4]/td/div/div[2]/div/div[1]/table/tbody/tr[2]/td/table/tbody/tr[2]/td/div/table/tbody/tr[2]/td[1]/span/table/tbody/tr/td')))
                ocDateField = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div/form/table[1]/tbody/tr[4]/td/div/div[2]/div/div[1]/table/tbody/tr[2]/td/table/tbody/tr[2]/td/div/table/tbody/tr[2]/td[1]/span/table/tbody/tr/td')
                ocDate = ocDateField.text
                sheet["D"+str(rowNumber)] = str(convert_date_with_map(str(ocDate)))
                workbook.save("############")
            except TimeoutException:
                print ("SAP failed, run script again...")


            time.sleep(1)
            
            original_window = driver.current_window_handle
            driver.switch_to.new_window("window")
            newWindow = driver.current_window_handle
            driver.switch_to.window(original_window)
            driver.close()
            driver.switch_to.window(newWindow)
            
            driver.get("edge://settings/clearBrowserData")
            driver.get("edge://settings/clearBrowserData")
            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div/div/div/div[2]/div/div[2]/div/div[2]/button[1]')))
                clearCache = driver.find_element(By.XPATH, '/html/body/div[2]/div/div/div/div[2]/div/div[2]/div/div[2]/button[1]')
                clearCache.click()
            except TimeoutException:
                print ("SAP failed, run script again...")
            time.sleep(1)
            driver.get("####################################")
            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/table/tbody/tr/td[1]/div/div/ul/span[1]/li/span/a')))
                inicioButton = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[1]/td/table/tbody/tr[3]/td/table/tbody/tr/td[1]/div/div/ul/span[1]/li/span/a')
                inicioButton.click()
            except TimeoutException:
                print ("SAP failed, run script again...")
            time.sleep(2)

            driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[3]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr/td[2]/div[2]/table/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td/div[1]/div/div[8]/div/div/div/div[2]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/a')))
                seeMoreButton = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[3]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr/td[2]/div[2]/table/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td/div[1]/div/div[8]/div/div/div/div[2]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/a')
                seeMoreButton.click()
            except TimeoutException:
                print ("SAP failed, run script again...")


            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/div[2]/div/a[1]')))
                ocButton = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/div[2]/div/a[1]')
                ocButton.click()
            except TimeoutException:
                print ("SAP failed, run script again...")
            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[2]/tbody/tr[1]/td[1]/div/table[2]/tbody/tr[18]/td[2]/table/tbody/tr[2]/td[2]/nobr/div[2]/div[1]/input')))
                sinceSearch = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[2]/tbody/tr[1]/td[1]/div/table[2]/tbody/tr[18]/td[2]/table/tbody/tr[2]/td[2]/nobr/div[2]/div[1]/input')
                sinceSearch.clear()
                time.sleep(1)
                sinceSearch.send_keys(str(get_date_30_days_before()))
            except TimeoutException:
                print ("SAP failed, run script again...")





            time.sleep(1)  
            driver.execute_script("window.scrollTo(5,document.body.scrollHeight)")
            driver.execute_script("window.scrollTo(5,document.body.scrollHeight)")
            
            try:    
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[2]/tbody/tr[1]/td[1]/div/table[2]/tbody/tr[15]/td[2]/div/div')))
                stateOptions = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[2]/tbody/tr[1]/td[1]/div/table[2]/tbody/tr[15]/td[2]/div/div')
                stateOptions.click()
            except TimeoutException:
                print ("SAP failed, run script again...")


            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[2]/tbody/tr[1]/td[1]/div/table[2]/tbody/tr[15]/td[2]/div/div/div/div[10]')))
                pedidoButton = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[2]/tbody/tr[1]/td[1]/div/table[2]/tbody/tr[15]/td[2]/div/div/div/div[10]')
                pedidoButton.click()
            except TimeoutException:
                print ("SAP failed, run script again...")





            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[3]/tbody/tr/td[2]/table/tbody/tr/td[1]/button')))
                searchButton = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[1]/td/div[1]/table[3]/tbody/tr/td[2]/table/tbody/tr/td[1]/button')
                searchButton.click()
            except TimeoutException:
                print ("SAP failed, run script again...")
            time.sleep(1)



            try: 
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[2]/td/div[1]/table/tbody/tr[2]/td/table/tbody/tr[2]/td/div/table/tbody/tr['+str(i)+']/td[2]')))
                searchButton = driver.find_element(By.XPATH, '/html/body/div[5]/div[2]/table/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/table/tbody/tr[3]/td[2]/div/div/form/table/tbody/tr[2]/td/div[1]/table/tbody/tr[2]/td/table/tbody/tr[2]/td/div/table/tbody/tr['+str(i)+']/td[2]')
                something = searchButton.text
            except TimeoutException:
                print ("SAP failed, run script again...")


    except TimeoutException or NoSuchElementException:
        break
    i += 1

time.sleep(2)